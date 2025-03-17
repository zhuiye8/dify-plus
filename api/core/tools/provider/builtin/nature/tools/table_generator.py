"""
表格生成工具实现
"""
from typing import Any, Union, List, Dict
import json
import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from core.tools.entities.tool_entities import ToolInvokeMessage
from core.tools.tool.builtin_tool import BuiltinTool


class TableGeneratorTool(BuiltinTool):
    """
    表格生成工具类
    
    从结构化数据生成表格文件，支持 JSON、CSV 和文本格式
    """

    def _parse_text_data(self, text: str) -> pd.DataFrame:
        """解析文本格式的问答数据为 DataFrame"""
        print("开始解析文本数据...")
        print("接收到的文本数据:", text)
        
        # 处理转义的换行符
        text = text.replace('\\n', '\n')
        
        # 按空行分割问答对
        qa_pairs = [pair.strip() for pair in text.split('\n\n') if pair.strip()]
        print("分割后的问答对:", qa_pairs)
        
        questions = []
        answers = []
        
        for pair in qa_pairs:
            # 分割问题和答案
            lines = pair.split('\n')
            print("处理问答对:", lines)
            if len(lines) >= 2:
                # 移除问题号和答案号
                question = lines[0].strip().split('：', 1)[1] if '：' in lines[0] else lines[0].strip()
                answer = lines[1].strip().split('：', 1)[1] if '：' in lines[1] else lines[1].strip()
                questions.append(question)
                answers.append(answer)
        
        print("提取的问题:", questions)
        print("提取的答案:", answers)
        
        if not questions or not answers:
            raise ValueError("未能从文本中提取到有效的问答对")
            
        df = pd.DataFrame({
            '问题': questions,
            '答案': answers
        })
        print("生成的DataFrame:", df)
        return df

    def _parse_input_data(self, data_str: str) -> pd.DataFrame:
        """解析输入数据为 DataFrame"""
        print("开始解析输入数据...")
        print("接收到的数据:", data_str)
        
        try:
            # 尝试解析为 JSON
            data = json.loads(data_str)
            print("JSON解析结果:", data)
            
            if isinstance(data, list):
                return pd.DataFrame(data)
            elif isinstance(data, dict):
                if 'data' in data:
                    print("检测到data字段，开始解析文本数据")
                    return self._parse_text_data(data['data'])
                else:
                    return pd.DataFrame([data])
            else:
                raise ValueError("JSON 数据必须是对象数组或字典")
        except json.JSONDecodeError:
            print("JSON解析失败，尝试其他格式")
            try:
                # 尝试解析为 CSV
                df = pd.read_csv(io.StringIO(data_str))
                # 如果 CSV 解析后 DataFrame 为空，则使用文本解析
                if df.empty:
                    print("CSV解析结果为空，尝试文本解析")
                    return self._parse_text_data(data_str)
                return df
            except Exception:
                try:
                    return self._parse_text_data(data_str)
                except Exception as e:
                    raise ValueError(f"无法解析数据，既不是有效的 JSON、CSV 也不是有效的文本格式: {str(e)}")

    def _create_table_from_data(self, df: pd.DataFrame, sheet_name: str, include_styling: bool) -> bytes:
        """从 DataFrame 创建表格文件"""
        print("开始创建表格...")
        print("DataFrame内容:", df)
        
        if df.empty:
            raise ValueError("没有数据可以生成表格文件")
            
        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        
        # 设置工作表名称
        if sheet_name:
            ws.title = sheet_name
            
        # 添加数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            ws.append(row)
            
            # 应用样式
            if include_styling:
                if r_idx == 0:  # 标题行
                    for c_idx, cell in enumerate(ws[1], 1):
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = Border(
                            bottom=Side(style='thin'),
                            top=Side(style='thin'),
                            left=Side(style='thin'),
                            right=Side(style='thin')
                        )
                else:  # 数据行
                    for c_idx, cell in enumerate(ws[r_idx + 1], 1):
                        cell.alignment = Alignment(horizontal='left', wrap_text=True)  # 添加自动换行
                        cell.border = Border(
                            bottom=Side(style='thin'),
                            top=Side(style='thin'),
                            left=Side(style='thin'),
                            right=Side(style='thin')
                        )
                        
                        # 交替行颜色
                        if r_idx % 2 == 1:
                            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # 调整列宽
        for i, column in enumerate(df.columns, 1):
            column_width = max(len(str(column)), df[column].astype(str).map(len).max() if len(df) > 0 else 0)
            ws.column_dimensions[get_column_letter(i)].width = min(max(column_width + 2, 10), 100)  # 增加最大列宽
            
        # 调整行高
        for row in ws.rows:
            ws.row_dimensions[row[0].row].height = 60  # 设置固定行高
            
        # 保存到内存
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer.getvalue()

    def _invoke(
        self,
        user_id: str,
        tool_parameters: dict[str, Any],
    ) -> Union[ToolInvokeMessage, list[ToolInvokeMessage]]:
        """
        执行表格生成逻辑
        """
        print("开始执行表格生成...")
        print("接收到的参数:", tool_parameters)
        
        data_str = tool_parameters.get("data")
        sheet_name = tool_parameters.get("sheet_name", "Sheet1")
        include_styling = tool_parameters.get("include_styling", True)

        # 校验输入
        if not data_str:
            raise Exception("数据未提供")

        try:
            # 解析数据
            df = self._parse_input_data(data_str)
            
            # 创建表格文件
            table_content = self._create_table_from_data(df, sheet_name, include_styling)
            
            return [
                self.create_text_message("表格文件生成成功"),
                self.create_blob_message(
                    blob=table_content,
                    meta={
                        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "name": "generated_table.xlsx"
                    },
                    save_as=self.VariableKey.DOCUMENT
                )
            ]
            
        except Exception as e:
            print("发生错误:", str(e))
            raise Exception(f"生成表格文件失败: {str(e)}")

    def validate_credentials(self, credentials: dict[str, Any], tool_parameters: dict[str, Any]) -> None:
        """验证参数"""
        # 检查必要参数是否存在
        if "data" not in tool_parameters:
            raise Exception("数据参数缺失")

        # 验证数据格式
        data_str = tool_parameters.get("data")
        if isinstance(data_str, str):
            try:
                # 尝试解析为 JSON
                json.loads(data_str)
            except json.JSONDecodeError:
                try:
                    # 尝试解析为 CSV
                    pd.read_csv(io.StringIO(data_str))
                except Exception:
                    raise Exception("数据格式错误，必须为有效的 JSON 或 CSV") 