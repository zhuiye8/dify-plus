"""
Excel 生成工具实现
"""
from typing import Any, Union, List, Dict
import json
import io
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from core.tools.entities.tool_entities import ToolInvokeMessage
from core.tools.tool.builtin_tool import BuiltinTool


class ExcelGeneratorTool(BuiltinTool):
    """
    Excel 生成工具类
    
    从结构化数据生成 Excel 文件，支持 JSON 和 CSV 格式
    """

    def _parse_input_data(self, data_str: str) -> pd.DataFrame:
        """解析输入数据为 DataFrame"""
        try:
            # 尝试解析为 JSON
            data = json.loads(data_str)
            if isinstance(data, list):
                return pd.DataFrame(data)
            else:
                raise ValueError("JSON 数据必须是对象数组")
        except json.JSONDecodeError:
            try:
                # 尝试解析为 CSV
                return pd.read_csv(io.StringIO(data_str))
            except Exception as e:
                raise ValueError(f"无法解析数据，既不是有效的 JSON 也不是有效的 CSV: {str(e)}")

    def _create_excel_from_data(self, df: pd.DataFrame, sheet_name: str, include_styling: bool) -> bytes:
        """从 DataFrame 创建 Excel 文件"""
        if df.empty:
            raise ValueError("没有数据可以生成 Excel 文件")
            
        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        
        # 设置工作表名称
        if sheet_name:
            ws.title = sheet_name
            
        # 添加数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            ws.append(row)
            
            # 应用样式
            if include_styling:
                if r_idx == 0:  # 标题行
                    for c_idx, cell in enumerate(ws[1], 1):
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = Border(
                            bottom=Side(style='thin'),
                            top=Side(style='thin'),
                            left=Side(style='thin'),
                            right=Side(style='thin')
                        )
                else:  # 数据行
                    for c_idx, cell in enumerate(ws[r_idx + 1], 1):
                        cell.alignment = Alignment(horizontal='left')
                        cell.border = Border(
                            bottom=Side(style='thin'),
                            top=Side(style='thin'),
                            left=Side(style='thin'),
                            right=Side(style='thin')
                        )
                        
                        # 交替行颜色
                        if r_idx % 2 == 1:
                            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # 调整列宽
        for i, column in enumerate(df.columns, 1):
            column_width = max(len(str(column)), df[column].astype(str).map(len).max() if len(df) > 0 else 0)
            ws.column_dimensions[get_column_letter(i)].width = min(max(column_width + 2, 10), 50)
            
        # 保存到内存
        output_buffer = io.BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer.getvalue()

    def _invoke(
        self,
        user_id: str,
        tool_parameters: dict[str, Any],
    ) -> Union[ToolInvokeMessage, list[ToolInvokeMessage]]:
        """
        执行 Excel 生成逻辑
        """
        data_str = tool_parameters.get("data")
        sheet_name = tool_parameters.get("sheet_name", "Sheet1")
        include_styling = tool_parameters.get("include_styling", True)

        # 校验输入
        if not data_str:
            raise Exception("数据未提供")

        try:
            # 解析数据
            df = self._parse_input_data(data_str)
            
            # 创建 Excel 文件
            excel_content = self._create_excel_from_data(df, sheet_name, include_styling)
            
            return [
                self.create_text_message("Excel 文件生成成功"),
                self.create_blob_message(
                    blob=excel_content,
                    meta={
                        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "name": "generated_excel.xlsx"
                    },
                    save_as=self.VariableKey.DOCUMENT
                )
            ]
            
        except Exception as e:
            raise Exception(f"生成 Excel 文件失败: {str(e)}")

    def validate_credentials(self, credentials: dict[str, Any], tool_parameters: dict[str, Any]) -> None:
        """验证参数"""
        # 检查必要参数是否存在
        if "data" not in tool_parameters:
            raise Exception("数据参数缺失")

        # 验证数据格式
        data_str = tool_parameters.get("data")
        try:
            self._parse_input_data(data_str)
        except Exception as e:
            raise Exception(f"数据格式验证失败: {str(e)}") 