"""
Excel 生成工具实现
"""
from typing import Any, Union, ClassVar, Optional, List, Dict
import json
import requests
import re
import os
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from core.tools.entities.tool_entities import ToolInvokeMessage
from core.tools.tool.builtin_tool import BuiltinTool
from core.file.enums import FileType, FileAttribute
from core.file.file_manager import download, get_attr


class ExcelGeneratorTool(BuiltinTool):
    """
    Excel 生成工具类
    
    从 JSON 数据生成 Excel 文件，支持 LLM 增强的数据格式化和样式
    """
    DIFY_API_URL: ClassVar[str] = "https://api.dify.ai/v1"

    def _invoke_dify_llm(self, prompt: str, api_key: str) -> str:
        """调用 Dify 语言模型 API"""
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        data = {
            "messages": [{"role": "user", "content": prompt}],
            "response_mode": "blocking"
        }
        
        try:
            response = requests.post(f"{self.DIFY_API_URL}/chat-messages", headers=headers, json=data)
            response.raise_for_status()
            result = response.json()
            return result.get("answer", "").strip()
        except requests.exceptions.RequestException as e:
            raise Exception(f"语言模型调用失败: {str(e)}")
        except json.JSONDecodeError:
            raise Exception(f"语言模型返回的JSON格式无效: {response.text}")

    def _format_data_with_llm(self, data: Any, format_instructions: str, api_key: str) -> Dict[str, Any]:
        """使用 LLM 格式化数据为适合 Excel 的结构"""
        # 准备提示
        prompt = (
            f"我需要将以下 JSON 数据转换为适合 Excel 表格的格式。\n\n"
            f"数据:\n{json.dumps(data, ensure_ascii=False, indent=2)}\n\n"
        )
        
        if format_instructions:
            prompt += f"特殊格式要求: {format_instructions}\n\n"
            
        prompt += (
            "请将数据转换为一个扁平化的对象数组，其中每个对象代表一行，每个键代表一列。\n"
            "如果数据已经是数组格式，请优化列名和结构以便于在 Excel 中查看。\n"
            "如果数据是嵌套的，请将其扁平化，并为嵌套字段创建有意义的列名。\n"
            "请以 JSON 格式返回结果，确保它是一个对象数组，每个对象代表一行数据。\n"
            "同时，请提供一个建议的列顺序数组，按照逻辑顺序排列列名。\n\n"
            "返回格式:\n"
            "{\n"
            "  \"data\": [{行1}, {行2}, ...],\n"
            "  \"columns\": [\"列1\", \"列2\", ...],\n"
            "  \"title\": \"建议的表格标题\"\n"
            "}"
        )
        
        # 调用 LLM
        response = self._invoke_dify_llm(prompt, api_key)
        
        # 解析 LLM 返回的 JSON
        try:
            # 提取 JSON 部分
            json_match = re.search(r'```json\s*([\s\S]*?)\s*```', response)
            if json_match:
                json_str = json_match.group(1)
            else:
                json_str = response
                
            result = json.loads(json_str)
            
            # 验证结果格式
            if not isinstance(result, dict):
                raise ValueError("LLM 返回的结果不是有效的 JSON 对象")
                
            if "data" not in result or not isinstance(result["data"], list):
                raise ValueError("LLM 返回的结果缺少 'data' 数组")
                
            return result
            
        except json.JSONDecodeError as e:
            raise Exception(f"无法解析 LLM 返回的 JSON: {str(e)}\n原始响应: {response}")
        except Exception as e:
            raise Exception(f"处理 LLM 响应时出错: {str(e)}")

    def _create_excel_from_data(self, formatted_data: Dict[str, Any], sheet_name: str, include_styling: bool) -> bytes:
        """从格式化数据创建 Excel 文件"""
        # 提取数据和列信息
        data = formatted_data.get("data", [])
        columns = formatted_data.get("columns", [])
        title = formatted_data.get("title", "")
        
        if not data:
            raise ValueError("没有数据可以生成 Excel 文件")
            
        # 如果没有指定列顺序，使用第一行数据的键作为列
        if not columns and data:
            columns = list(data[0].keys())
            
        # 创建 DataFrame
        df = pd.DataFrame(data)
        
        # 如果有指定列顺序，重新排列列
        if columns:
            # 只使用存在于 DataFrame 中的列
            valid_columns = [col for col in columns if col in df.columns]
            # 添加任何在 DataFrame 中但不在 columns 中的列
            missing_columns = [col for col in df.columns if col not in valid_columns]
            ordered_columns = valid_columns + missing_columns
            df = df[ordered_columns]
            
        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        
        # 设置工作表名称
        if sheet_name:
            ws.title = sheet_name
            
        # 如果有标题，添加标题行
        if title and include_styling:
            ws.append([title])
            ws.merge_cells(f'A1:{get_column_letter(len(df.columns))}1')
            title_cell = ws['A1']
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')
            
            # 添加空行
            ws.append([])
            
            # 调整起始行
            start_row = 3
        else:
            start_row = 1
            
        # 添加数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            ws.append(row)
            
            # 应用样式
            if include_styling:
                if r_idx == 0:  # 标题行
                    for c_idx, cell in enumerate(ws[start_row], 1):
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = Border(
                            bottom=Side(style='thin'),
                            top=Side(style='thin'),
                            left=Side(style='thin'),
                            right=Side(style='thin')
                        )
                else:  # 数据行
                    for c_idx, cell in enumerate(ws[start_row + r_idx], 1):
                        cell.alignment = Alignment(horizontal='left')
                        cell.border = Border(
                            bottom=Side(style='thin'),
                            top=Side(style='thin'),
                            left=Side(style='thin'),
                            right=Side(style='thin')
                        )
                        
                        # 交替行颜色
                        if r_idx % 2 == 1:
                            cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # 调整列宽
        for i, column in enumerate(df.columns, 1):
            column_width = max(len(str(column)), df[column].astype(str).map(len).max() if len(df) > 0 else 0)
            ws.column_dimensions[get_column_letter(i)].width = min(max(column_width + 2, 10), 50)
            
        # 保存到内存
        output_buffer = BytesIO()
        wb.save(output_buffer)
        output_buffer.seek(0)
        return output_buffer.getvalue()

    def _invoke(
        self,
        user_id: str,
        tool_parameters: dict[str, Any],
    ) -> Union[ToolInvokeMessage, list[ToolInvokeMessage]]:
        """
        执行 Excel 生成逻辑
        """
        # 获取凭证和参数
        dify_api_key = self.runtime.credentials.get("dify_api_key")
        if not dify_api_key:
            raise Exception("Dify API Key 未提供")

        data_json = tool_parameters.get("data")
        sheet_name = tool_parameters.get("sheet_name", "Sheet1")
        format_instructions = tool_parameters.get("format_instructions", "")
        include_styling = tool_parameters.get("include_styling", True)

        # 校验输入
        if not data_json:
            raise Exception("数据未提供")

        # 解析 JSON 数据
        try:
            data = json.loads(data_json) if isinstance(data_json, str) else data_json
        except json.JSONDecodeError:
            raise Exception("数据格式错误，必须为有效的 JSON")

        try:
            # 使用 LLM 格式化数据
            formatted_data = self._format_data_with_llm(data, format_instructions, dify_api_key)
            
            # 创建 Excel 文件
            excel_content = self._create_excel_from_data(formatted_data, sheet_name, include_styling)
            
            # 生成文件名
            output_filename = "generated_excel.xlsx"
            
            return [
                self.create_text_message("Excel 文件生成成功"),
                self.create_blob_message(
                    blob=excel_content,
                    meta={
                        "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "name": output_filename
                    },
                    save_as=self.VariableKey.DOCUMENT
                )
            ]
            
        except Exception as e:
            raise Exception(f"生成 Excel 文件失败: {str(e)}")

    def validate_credentials(self, credentials: dict[str, Any], tool_parameters: dict[str, Any]) -> None:
        """验证凭证和参数"""
        dify_api_key = credentials.get("dify_api_key")
        if not dify_api_key:
            raise Exception("Dify API Key 未提供")

        # 检查必要参数是否存在
        if "data" not in tool_parameters:
            raise Exception("数据参数缺失")

        # 验证JSON数据格式
        data_json = tool_parameters.get("data")
        if isinstance(data_json, str):
            try:
                json.loads(data_json)
            except json.JSONDecodeError:
                raise Exception("数据格式错误，必须为有效的 JSON")

        # 测试语言模型调用
        try:
            self._invoke_dify_llm("测试提示", dify_api_key)
        except Exception as e:
            raise Exception(f"语言模型调用测试失败: {str(e)}") 